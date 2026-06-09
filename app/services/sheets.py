import json
import os
from pathlib import Path
from typing import Any

from google.oauth2 import service_account
from googleapiclient.discovery import build
from openpyxl import load_workbook


INVENTORY_COLUMNS = [
    "제조사",
    "품명",
    "규격",
    "신품/중고",
    "무자료가격",
    "계산서가격",
    "입고일자",
    "제조사 판매가",
    "재고수량",
]

SAMPLE_ITEMS = [
    {
        "제조사": "티케이",
        "품명": "모터",
        "규격": "삼성 220V 1HP",
        "신품/중고": "신품",
        "무자료가격": "120000",
        "계산서가격": "132000",
        "입고일자": "2026-06-01",
        "제조사 판매가": "150000",
        "재고수량": "8",
    },
    {
        "제조사": "현대",
        "품명": "펌프",
        "규격": "LG 산업용 2HP",
        "신품/중고": "중고",
        "무자료가격": "90000",
        "계산서가격": "99000",
        "입고일자": "2026-05-18",
        "제조사 판매가": "130000",
        "재고수량": "3",
    },
    {
        "제조사": "오티스",
        "품명": "인버터",
        "규격": "LS 3상 380V",
        "신품/중고": "신품",
        "무자료가격": "210000",
        "계산서가격": "231000",
        "입고일자": "2026-05-30",
        "제조사 판매가": "260000",
        "재고수량": "5",
    },
    {
        "제조사": "기타",
        "품명": "센서",
        "규격": "근접센서 24V",
        "신품/중고": "중고",
        "무자료가격": "15000",
        "계산서가격": "16500",
        "입고일자": "2026-05-10",
        "제조사 판매가": "22000",
        "재고수량": "0",
    },
]


class InventoryStore:
    def __init__(self) -> None:
        self._cache: list[dict[str, str]] | None = None
        self._source = "none"

    def get_inventory(self, refresh: bool = False) -> dict[str, Any]:
        if refresh or self._cache is None:
            self._cache, self._source = self._load_inventory()

        return {
            "source": self._source,
            "items": self._cache,
        }

    def _load_inventory(self) -> tuple[list[dict[str, str]], str]:
        try:
            rows = self._load_sheet_rows()
            if rows:
                return rows, "google_sheets"
        except Exception:
            pass
        return SAMPLE_ITEMS, "sample"

    def _load_sheet_rows(self) -> list[dict[str, str]]:
        sheet_id = os.getenv("GOOGLE_SHEET_ID")
        account_json = os.getenv("GOOGLE_SERVICE_ACCOUNT_JSON")
        if not sheet_id or not account_json:
            return []

        service = self._build_service(["https://www.googleapis.com/auth/spreadsheets.readonly"])
        sheet_range = os.getenv("GOOGLE_SHEET_RANGE", "A:K")
        result = (
            service.spreadsheets()
            .values()
            .get(spreadsheetId=sheet_id, range=sheet_range)
            .execute()
        )
        values = result.get("values", [])
        if not values:
            return []

        headers = [str(value).strip() for value in values[0]]
        rows = []
        for value_row in values[1:]:
            item = {}
            for column in INVENTORY_COLUMNS:
                if column in headers:
                    index = headers.index(column)
                    item[column] = str(value_row[index]) if index < len(value_row) else ""
                else:
                    item[column] = ""
            rows.append(item)
        return rows

    def upload_xlsx(self, file_path: Path) -> dict[str, Any]:
        sheet_id = os.getenv("GOOGLE_SHEET_ID")
        if not sheet_id:
            raise ValueError("GOOGLE_SHEET_ID is required")

        values = self._read_xlsx_values(file_path)
        service = self._build_service(["https://www.googleapis.com/auth/spreadsheets"])
        sheet_range = os.getenv("GOOGLE_SHEET_RANGE", "A:K")
        service.spreadsheets().values().clear(
            spreadsheetId=sheet_id,
            range=sheet_range,
            body={},
        ).execute()
        service.spreadsheets().values().update(
            spreadsheetId=sheet_id,
            range=sheet_range,
            valueInputOption="USER_ENTERED",
            body={"values": values},
        ).execute()
        self._cache = None
        return {"rows": max(len(values) - 1, 0)}

    def _read_xlsx_values(self, file_path: Path) -> list[list[str]]:
        workbook = load_workbook(file_path, data_only=True)
        sheet = workbook.active
        values: list[list[str]] = []
        for row in sheet.iter_rows(values_only=True):
            values.append(["" if value is None else str(value) for value in row])

        if not values:
            raise ValueError("Uploaded XLSX has no rows")

        headers = [value.strip() for value in values[0]]
        missing_columns = [column for column in INVENTORY_COLUMNS if column not in headers]
        if missing_columns:
            raise ValueError(f"Missing columns: {', '.join(missing_columns)}")
        return values

    def _build_service(self, scopes: list[str]):
        account_json = os.getenv("GOOGLE_SERVICE_ACCOUNT_JSON")
        if not account_json:
            raise ValueError("GOOGLE_SERVICE_ACCOUNT_JSON is required")

        account_info = self._load_service_account_info(account_json)
        credentials = service_account.Credentials.from_service_account_info(
            account_info,
            scopes=scopes,
        )
        return build("sheets", "v4", credentials=credentials)

    def _load_service_account_info(self, account_json: str) -> dict[str, Any]:
        try:
            return json.loads(account_json)
        except json.JSONDecodeError:
            path = Path(account_json)
            if path.exists():
                return json.loads(path.read_text(encoding="utf-8"))
            raise
